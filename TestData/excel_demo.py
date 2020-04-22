# install pyxl
# > pip3 install openpyxl

import sys
import openpyxl

book = openpyxl.load_workbook(sys.path[0] + '/demo_excel.xlsx')
sheet = book.active
cell_2_2 = sheet.cell(row=2, column=2).value

print(sheet.max_column)
print(sheet.max_row)
print(sheet['A2'].value)

print('----------------')
dc = [{}]


def getHeaders():
    return [sheet.cell(row=1, column=i).value for i in range(2, sheet.max_column + 1)]


def getAllDataAsListOfDict():
    lst = []
    for i in range(2, sheet.max_row + 1):  # to get rows
        dc = {}
        for j in range(2, sheet.max_column + 1):
            dc[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        lst.append(dc)

    print(lst)


getAllDataAsListOfDict()
