import sys
import openpyxl


class HomePageData:
    test_homePage_data = [{'name': 'John', 'email': 'doe@asd.com', 'gender': 'Male'},
                          {'name': 'Janet', 'email': "fwe@cm.tr", 'gender': "Female"}]

    @staticmethod
    def getAllDataAsListOfDict():
        # /Users/mac/PycharmProjects/PythonSeleniumFramework/TestData/demo_excel.xlsx
        book = openpyxl.load_workbook(sys.path[0] + '/TestData/demo_excel.xlsx')
        sheet = book.active
        ls = []
        for i in range(2, sheet.max_row + 1):  # to get rows
            dc = {}
            for j in range(2, sheet.max_column + 1):
                dc[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            ls.append(dc)

        print(ls)
        return ls
